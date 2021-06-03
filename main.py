from openpyxl.styles.borders import Border, Side
import openpyxl as xl

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

wb = xl.load_workbook('numbers.xlsx')
sheet = wb['Sheet1']
sheet2 = wb.create_sheet('modified')
new_dict = {}
row_count = sheet.max_row
column_count = sheet.max_column
counter = 1

move_position_row = 1
move_position_column = 5

number_list = []
index = 0

for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        number_list.append(sheet.cell(row=row, column=column).value)

for i in range(move_position_row, move_position_row + sheet.max_row + 1):
    for j in range(move_position_column, move_position_column + sheet.max_column):
        if index == len(number_list):
            break
        sheet2.cell(row=i, column=j).value = number_list[index]
        if i != move_position_row and j != move_position_column:
            sheet2.cell(row=i, column=j).border = thin_border
        index += 1

for row in range(1, sheet.max_row + 1):
    blank_list = []
    for column in range(1, sheet.max_column + 1):
        blank_list.append(sheet.cell(row=row, column=column).value)
    new_dict[row] = blank_list

#if you wish only outline borders - use this code below

# for row in range(1 + move_position_row, move_position_row + sheet.max_row):
#     for key in new_dict.keys():
#         for number in new_dict[key]:
#             for column in range(move_position_column + 1, move_position_column + sheet.max_column):
#                 if key == 1 and row == move_position_row + 1:
#                     if column == move_position_column + 1:
#                         sheet2.cell(row=row, column=column).border = Border(left=Side(style='thin'),
#                                                                             top=Side(style='thin'))
#                     elif column == move_position_column + sheet.max_column - 1:
#                         sheet2.cell(row=row, column=column).border = Border(right=Side(style='thin'),
#                                                                             top=Side(style='thin'))
#                     else:
#                         sheet2.cell(row=row, column=column).border = Border(top=Side(style='thin'))
#
#                 elif key == row_count and row == sheet.max_row:
#                     if column == move_position_column + 1:
#                         sheet2.cell(row=row, column=column).border = Border(left=Side(style='thin'),
#                                                                             bottom=Side(style='thin'))
#                     elif column == move_position_column + sheet.max_column - 1:
#                         sheet2.cell(row=row, column=column).border = Border(right=Side(style='thin'),
#                                                                             bottom=Side(style='thin'))
#                     else:
#                         sheet2.cell(row=row, column=column).border = Border(bottom=Side(style='thin'))
#
#                 elif column == move_position_column + 1 and row != move_position_row + 1:
#                     sheet2.cell(row=row, column=column).border = Border(left=Side(style='thin'))
#                 elif column == move_position_column + sheet.max_column - 1 and row != move_position_row + 1:
#                     sheet2.cell(row=row, column=column).border = Border(right=Side(style='thin'))

wb.save('modified.xlsx')
